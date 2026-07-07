import openpyxl
from datetime import datetime, date
from sqlalchemy.orm import Session
from . import models


def get_sheet_names(file_path: str):
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception as e:
        print(f"Error reading sheets: {e}")
        raise ValueError(f"Could not read sheets: {str(e)}")


def parse_tracking_excel(file_path: str, db: Session, sheet_name: str = "Summary", model_name: str = "603"):
    """Parse tracking Excel with dynamic column detection."""

    # Load workbook twice: once for values, once for font colors
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    wb_fmt = openpyxl.load_workbook(file_path, data_only=False)

    if sheet_name not in wb_data.sheetnames:
        wb_data.close()
        wb_fmt.close()
        raise ValueError(f"Sheet '{sheet_name}' not found.")

    sheet_data = wb_data[sheet_name]
    sheet_fmt = wb_fmt[sheet_name]

    # Build 2D arrays
    data_rows = []
    for row in sheet_data.iter_rows():
        data_rows.append([cell.value for cell in row])

    fmt_rows = []
    for row in sheet_fmt.iter_rows():
        fmt_data = []
        for cell in row:
            is_red = False
            is_black = False
            try:
                # Check font color (for alerts)
                if cell.font and cell.font.color and cell.font.color.rgb:
                    rgb = str(cell.font.color.rgb).upper()
                    if 'FF0000' in rgb:
                        is_red = True
                
                # Check background fill (to skip unreported/blacked-out cells)
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    fill_rgb = str(cell.fill.fgColor.rgb).upper()
                    # FF000000 is solid black. 00000000 is transparent/no-fill, do NOT skip transparent.
                    if fill_rgb in ('FF000000', '000000', 'FF000001'):
                        is_black = True
            except:
                pass
            fmt_data.append({"is_red": is_red, "is_black": is_black})
        fmt_rows.append(fmt_data)

    wb_data.close()
    wb_fmt.close()

    total_rows = len(data_rows)
    num_cols = max((len(r) for r in data_rows), default=0)

    # Debug log
    log = []
    log.append(f"=== PARSE START: {sheet_name} | {total_rows} rows x {num_cols} cols ===")

    # Category keywords
    cat_keywords = {
        "CUTTING": "CUTTING + PREPARATION",
        "COMPUTER STITCHING": "COMPUTER STITCHING",
        "SEWING": "SEWING",
        "ASSEMBLY": "ASSEMBLY",
    }

    def s(val):
        """Safe string."""
        return str(val).strip() if val is not None else ""

    def find_date(val):
        """Try to parse a date from various formats."""
        if val is None:
            return None
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        text = s(val)
        if '/' in text:
            try:
                parts = text.split('/')
                month, day = int(parts[0]), int(parts[1])
                return date(datetime.now().year, month, day)
            except:
                pass
        # Could be a number (Excel serial date)
        if isinstance(val, (int, float)) and val > 40000:
            try:
                from datetime import timedelta
                base = date(1899, 12, 30)
                return base + timedelta(days=int(val))
            except:
                pass
        return None

    records_added = 0
    seen_keys = set()  # Track (category, cell, date) to prevent duplicates
    row_idx = 0

    while row_idx < total_rows:
        row = data_rows[row_idx]
        # Check ALL columns in this row for a category keyword
        row_text = " ".join([s(v).upper() for v in row[:10] if v is not None])

        found_cat = None
        for kw, full_name in cat_keywords.items():
            if kw in row_text:
                found_cat = full_name
                break

        if not found_cat:
            row_idx += 1
            continue

        log.append(f"\n[CAT] '{found_cat}' at Excel row {row_idx + 1}")

        # --- DYNAMIC COLUMN DETECTION ---
        # Search for the "Cell" header in the rows below this category title.
        # Find WHICH COLUMN contains "Cell" to determine the layout.
        cell_col = -1  # column index where "Cell" header is
        header_idx = -1  # row index of the header

        for i in range(row_idx, min(row_idx + 15, total_rows)):
            r = data_rows[i]
            for c_idx, val in enumerate(r):
                if s(val).upper() == "CELL":
                    cell_col = c_idx
                    header_idx = i
                    break
            if cell_col >= 0:
                break

        if cell_col < 0:
            log.append(f"  [SKIP] No 'Cell' header found")
            row_idx += 1
            continue

        log.append(f"  [HDR] 'Cell' at col {cell_col} (col letter {chr(65+cell_col)}), row {header_idx + 1}")

        # Column layout relative to cell_col:
        # cell_col+0 = Cell name
        # cell_col+1 = Working Period
        # cell_col+2 = STD MP
        # cell_col+3 = ACT MP
        # cell_col+4 = Gap
        # cell_col+5 = Output (Output/Day, Output/H label)
        # cell_col+6 onwards = date value columns
        col_wp = cell_col + 1
        col_std = cell_col + 2
        col_act = cell_col + 3
        col_gap = cell_col + 4
        col_output_label = cell_col + 5
        col_date_start = cell_col + 6

        # Find the date row: scan for "/" in cells from col_date_start onwards
        date_row_idx = -1
        for i in range(header_idx + 1, min(header_idx + 5, total_rows)):
            r = data_rows[i]
            for c in range(col_date_start, len(r)):
                if find_date(r[c]) is not None:
                    date_row_idx = i
                    break
            if date_row_idx >= 0:
                break

        if date_row_idx < 0:
            log.append(f"  [SKIP] No date row found")
            row_idx += 1
            continue

        # Collect dates
        dates = {}
        dr = data_rows[date_row_idx]
        for c in range(col_date_start, len(dr)):
            d = find_date(dr[c])
            if d:
                dates[c] = d

        log.append(f"  [DATE] row {date_row_idx + 1}")

        # First pass: detect the correct year from datetime cells
        detected_year = None
        dr = data_rows[date_row_idx]
        for c in range(col_date_start, len(dr)):
            val = dr[c]
            if isinstance(val, datetime):
                detected_year = val.year
                break
            if isinstance(val, date):
                detected_year = val.year
                break
        
        if not detected_year:
            detected_year = datetime.now().year
            log.append(f"  [YEAR] No year detected in cells, defaulting to current year: {detected_year}")
        else:
            log.append(f"  [YEAR] Detected year from Excel data: {detected_year}")

        # Second pass: parse all dates using the detected year
        dates = {}
        for c in range(col_date_start, len(dr)):
            val = dr[c]
            d = None
            if isinstance(val, datetime):
                d = val.date()
            elif isinstance(val, date):
                d = val
            elif val is not None:
                text = s(val)
                if '/' in text:
                    try:
                        parts = text.split('/')
                        month, day = int(parts[0]), int(parts[1])
                        d = date(detected_year, month, day)
                    except:
                        pass
                elif isinstance(val, (int, float)) and val > 40000:
                    try:
                        from datetime import timedelta
                        base = date(1899, 12, 30)
                        d = base + timedelta(days=int(val))
                    except:
                        pass
            if d:
                dates[c] = d
        current = date_row_idx + 1
        while current < total_rows:
            r = data_rows[current]
            # Get cell name from the correct column
            cell_name_val = s(r[cell_col]) if cell_col < len(r) else ""
            output_label = s(r[col_output_label]).upper() if col_output_label < len(r) else ""

            # Check for end of this category block
            end_text = " ".join([s(v).upper() for v in r[:10] if v is not None])
            if any(kw in end_text for kw in cat_keywords) and "CELL" not in end_text:
                break
            if "TARGET" in end_text:
                break

            # Is this a Cell data row?
            if "CELL" in cell_name_val.upper():
                cell_name = cell_name_val
                wp = s(r[col_wp]) if col_wp < len(r) else ""
                try:
                    std = int(float(r[col_std])) if r[col_std] is not None else 0
                except:
                    std = 0
                try:
                    act = int(float(r[col_act])) if r[col_act] is not None else 0
                except:
                    act = 0
                try:
                    gap = int(float(r[col_gap])) if r[col_gap] is not None else 0
                except:
                    gap = 0

                # This row = Output/Day, next row = Output/H
                day_r = current
                h_r = current + 1 if current + 1 < total_rows else None

                log.append(f"    [CELL] '{cell_name}' row {current + 1}, wp={wp}, std={std}, act={act}, gap={gap}")

                for col, date_obj in dates.items():
                    try:
                        # Output/Day
                        dv = data_rows[day_r][col] if col < len(data_rows[day_r]) else None
                        out_day = float(dv) if dv is not None and s(dv) not in ('', '-') else 0.0

                        # Output/H
                        out_h = 0.0
                        if h_r and col < len(data_rows[h_r]):
                            hv = data_rows[h_r][col]
                            out_h = float(hv) if hv is not None and s(hv) not in ('', '-') else 0.0

                        # Font colors & Skip logic (Black background)
                        day_status = "normal"
                        hour_status = "normal"
                        
                        if day_r < len(fmt_rows) and col < len(fmt_rows[day_r]):
                            cel_fmt = fmt_rows[day_r][col]
                            if cel_fmt.get("is_black"):
                                log.append(f"      [SKIP] Col {col} is blacked out (unreported)")
                                continue
                            if cel_fmt.get("is_red"):
                                day_status = "alert"
                                
                        if h_r and h_r < len(fmt_rows) and col < len(fmt_rows[h_r]):
                            cel_fmt_h = fmt_rows[h_r][col]
                            if cel_fmt_h.get("is_black"):
                                # If hour row is black, we might still want day data, 
                                # but usually black means the whole date column is empty.
                                pass 
                            if cel_fmt_h.get("is_red"):
                                hour_status = "alert"

                        # Skip records with zero production (unreported placeholders)
                        if out_day == 0 and out_h == 0:
                            # log.append(f"      [SKIP] Col {col} has no production data (0/0)")
                            continue

                        # Skip records too far in the future (more than 31 days)
                        today = date.today()
                        if date_obj > today.replace(day=min(today.day + 31, 28)): # Rough 1 month check
                            # Better approach: 31 days from today
                            from datetime import timedelta
                            if date_obj > today + timedelta(days=31):
                                log.append(f"      [SKIP] Date {date_obj} is too far in the future.")
                                continue

                        # Upsert - skip if we already processed this key
                        rec_key = (found_cat, cell_name, date_obj)
                        if rec_key in seen_keys:
                            continue
                        seen_keys.add(rec_key)

                        existing = db.query(models.ProductionData).filter(
                            models.ProductionData.model_name == model_name,
                            models.ProductionData.category == found_cat,
                            models.ProductionData.cell == cell_name,
                            models.ProductionData.date == date_obj
                        ).first()

                        if existing:
                            existing.working_period = wp
                            existing.std_mp = std
                            existing.act_mp = act
                            existing.gap = gap
                            existing.output_day = out_day
                            existing.output_h = out_h
                            existing.day_status = day_status
                            existing.hour_status = hour_status
                        else:
                            db.add(models.ProductionData(
                                model_name=model_name,
                                category=found_cat, cell=cell_name,
                                working_period=wp, date=date_obj,
                                std_mp=std, act_mp=act, gap=gap,
                                output_day=out_day, output_h=out_h,
                                day_status=day_status, hour_status=hour_status
                            ))
                            db.flush()
                        records_added += 1
                    except Exception as e:
                        log.append(f"      [ERR] col {col}: {e}")
                        continue

                current += 2  # skip Output/H row
            elif cell_name_val.strip() == "":
                # Empty cell name — could be Output/H row or spacer
                if "OUTPUT" in output_label:
                    current += 1  # skip orphan output row
                else:
                    # Check if next row has a cell
                    if current + 1 < total_rows:
                        next_name = s(data_rows[current + 1][cell_col]) if cell_col < len(data_rows[current + 1]) else ""
                        if "CELL" in next_name.upper() or "BUFFER" in next_name.upper():
                            current += 1
                            continue
                    break
            else:
                current += 1

        row_idx = max(current, row_idx + 1)

    db.commit()
    log.append(f"\n=== DONE: {records_added} records ===")

    with open("parser_debug.log", "w", encoding="utf-8") as f:
        f.write("\n".join(log))

    return records_added


def parse_marketing_excel(file_path: str, db: Session, sheet_name: str = "Summary", model_name: str = "603"):
    """Parse marketing specific Excel format."""
    import re
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    
    sheet = wb[sheet_name]
    
    # Try to detect year from I2 or similar
    detected_year = datetime.now().year
    try:
        title_info = ""
        # Check a few potential cells for the date string like "Tuesday, April 28, 2026"
        for cell_ref in ['I2', 'H2', 'G2', 'J2', 'A2']:
            val = sheet[cell_ref].value
            if val:
                title_info = str(val)
                break
        
        year_match = re.search(r'20\d{2}', title_info)
        if year_match:
            detected_year = int(year_match.group(0))
    except:
        pass

    records_added = 0
    current_month_name = ""
    
    # Data starts from row 5 based on the screenshot
    # We scan to find the header first to be sure
    start_row = 5
    for r in range(1, 10):
        if str(sheet.cell(row=r, column=2).value).strip().lower() == 'date':
            start_row = r + 1
            break

    for row_idx in range(start_row, sheet.max_row + 1):
        # Column A: Month
        month_val = sheet.cell(row=row_idx, column=1).value
        if month_val:
            current_month_name = str(month_val).strip()
            
        # Column B: Date
        date_val = sheet.cell(row=row_idx, column=2).value
        
        if date_val is None:
            continue
        
        date_str = str(date_val).lower()
        if "total" in date_str or "grand" in date_str or not date_str.strip():
            continue
            
        # Try to construct the date
        actual_date = None
        try:
            if isinstance(date_val, (int, float)):
                day = int(date_val)
                month_map = {
                    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'may': 5, 'june': 6,
                    'july': 7, 'august': 8, 'september': 9, 'october': 10, 'november': 11, 'december': 12
                }
                month_num = month_map.get(current_month_name.lower(), 1)
                actual_date = date(detected_year, month_num, day)
            elif isinstance(date_val, (datetime, date)):
                actual_date = date_val.date() if hasattr(date_val, 'date') else date_val
            else:
                # Try parsing string
                try:
                    actual_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except:
                    continue
        except:
            continue

        if not actual_date:
            continue

        # Extract values
        pd_hrs = sheet.cell(row=row_idx, column=3).value
        target = sheet.cell(row=row_idx, column=4).value
        act_output = sheet.cell(row=row_idx, column=5).value
        total_cell = sheet.cell(row=row_idx, column=6).value
        gap = sheet.cell(row=row_idx, column=7).value
        act_vs_target = sheet.cell(row=row_idx, column=8).value
        remarks = sheet.cell(row=row_idx, column=9).value

        def safe_float(v):
            if v is None or str(v).strip() in ('', '-', '#DIV/0!'): return 0.0
            try: return float(v)
            except: return 0.0
            
        def safe_int(v):
            if v is None or str(v).strip() in ('', '-', '#DIV/0!'): return 0
            try: return int(float(v))
            except: return 0

        existing = db.query(models.MarketingData).filter(
            models.MarketingData.model_name == model_name,
            models.MarketingData.date == actual_date
        ).first()
        
        if existing:
            existing.month_name = current_month_name
            existing.pd_hrs = safe_float(pd_hrs)
            existing.target = safe_int(target)
            existing.act_output = safe_int(act_output)
            existing.total_cell = safe_int(total_cell)
            existing.gap = safe_int(gap)
            existing.act_vs_target = safe_float(act_vs_target)
            existing.remarks = str(remarks) if remarks else ""
        else:
            db.add(models.MarketingData(
                model_name=model_name,
                date=actual_date,
                month_name=current_month_name,
                pd_hrs=safe_float(pd_hrs),
                target=safe_int(target),
                act_output=safe_int(act_output),
                total_cell=safe_int(total_cell),
                gap=safe_int(gap),
                act_vs_target=safe_float(act_vs_target),
                remarks=str(remarks) if remarks else ""
            ))
        records_added += 1
    
    db.commit()
    wb.close()
    return records_added
