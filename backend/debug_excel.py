"""
Debug script: Dump the first 40 rows of the Summary sheet
to understand the exact Excel structure.
Run: python debug_excel.py path_to_file.xlsx
"""
import openpyxl, sys, glob, os

# Find the most recent uploaded temp file, or use arg
if len(sys.argv) > 1:
    fp = sys.argv[1]
else:
    # Check if there's a temp file from a recent upload
    temps = glob.glob("temp_*.xlsx") + glob.glob("temp_*.xls")
    if temps:
        fp = temps[0]
    else:
        print("Usage: python debug_excel.py <path_to_excel>")
        print("Or place a temp_*.xlsx file in this directory.")
        sys.exit(1)

print(f"Opening: {fp}")
wb = openpyxl.load_workbook(fp, data_only=True)
print(f"Sheets: {wb.sheetnames}")

sheet = wb[wb.sheetnames[0]]  # First sheet
print(f"\nUsing sheet: {wb.sheetnames[0]}")
print(f"Dimensions: {sheet.dimensions}")
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

rows = list(sheet.rows)
print(f"Total rows via list(sheet.rows): {len(rows)}")
print()

# Print first 50 rows, columns A-G (0-6) + first date columns
for i, row in enumerate(rows[:50]):
    vals = []
    for j, cell in enumerate(row[:30]):  # first 30 columns
        v = cell.value
        if v is not None:
            vals.append(f"[{j}]={repr(v)}")
    if vals:
        print(f"Row {i} (Excel {i+1}): {', '.join(vals)}")
    else:
        print(f"Row {i} (Excel {i+1}): (all empty)")

wb.close()
